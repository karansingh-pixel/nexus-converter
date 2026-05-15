#!/usr/bin/env python3
"""
NEXUS - Standard to Beta Converter  v2.0
Data + Comments (Karan Singh) + Colors + Summary Tab
Ops-Avengers | Pattern E-commerce | Built by Karan Singh
"""

from flask import Flask, request, send_file, jsonify
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import io, json, zipfile, re, uuid
from datetime import datetime

app = Flask(__name__)

# ── FULL MAPPING ──
MAPPING_BASE = {
'item_sku':'contribution_sku#1.value',
'update_delete':'::record_action',
'feed_product_type':'product_type#1.value',
'item_name':'item_name[marketplace_id=__MP__][language_tag=en_us]#1.value',
'brand_name':'brand[marketplace_id=__MP__][language_tag=en_us]#1.value',
'external_product_id_type':'amzn1.volt.ca.product_id_type',
'external_product_id':'amzn1.volt.ca.product_id_value',
'recommended_browse_nodes':'recommended_browse_nodes[marketplace_id=__MP__]#1.value',
'model_number':'model_number[marketplace_id=__MP__]#1.value',
'manufacturer':'manufacturer[marketplace_id=__MP__][language_tag=en_us]#1.value',
'condition_type':'condition_type[marketplace_id=__MP__]#1.value',
'list_price_with_tax':'list_price[marketplace_id=__MP__]#1.value_with_tax',
'list_price':'list_price[marketplace_id=__MP__]#1.value',
'product_tax_code':'product_tax_code#1.value',
'fulfillment_center_id':'fulfillment_availability#1.fulfillment_channel_code',
'standard_price':'purchasable_offer[marketplace_id=__MP__][audience=all]#1.our_price#1.schedule#1.value_with_tax',
'product_description':'product_description[marketplace_id=__MP__][language_tag=en_us]#1.value',
'bullet_point1':'bullet_point[marketplace_id=__MP__][language_tag=en_us]#1.value',
'bullet_point2':'bullet_point[marketplace_id=__MP__][language_tag=en_us]#2.value',
'bullet_point3':'bullet_point[marketplace_id=__MP__][language_tag=en_us]#3.value',
'bullet_point4':'bullet_point[marketplace_id=__MP__][language_tag=en_us]#4.value',
'bullet_point5':'bullet_point[marketplace_id=__MP__][language_tag=en_us]#5.value',
'lifestyle':'lifestyle[marketplace_id=__MP__][language_tag=en_us]#1.value',
'style_name':'style[marketplace_id=__MP__][language_tag=en_us]#1.value',
'department_name':'department[marketplace_id=__MP__][language_tag=en_us]#1.value',
'target_gender':'target_gender[marketplace_id=__MP__]#1.value',
'age_range_description':'age_range_description[marketplace_id=__MP__][language_tag=en_us]#1.value',
'apparel_size_system':'apparel_size[marketplace_id=__MP__]#1.size_system',
'apparel_size_class':'apparel_size[marketplace_id=__MP__]#1.size_class',
'apparel_size':'apparel_size[marketplace_id=__MP__]#1.size',
'apparel_size_to':'apparel_size[marketplace_id=__MP__]#1.size_to',
'apparel_size_body_type':'apparel_size[marketplace_id=__MP__]#1.body_type',
'material_type1':'material[marketplace_id=__MP__][language_tag=en_us]#1.value',
'fabric_type':'fabric_type[marketplace_id=__MP__][language_tag=en_us]#1.value',
'water_resistance_level':'water_resistance_level[marketplace_id=__MP__]#1.value',
'color_name':'color[marketplace_id=__MP__][language_tag=en_us]#1.standardized_values#1',
'color_map':'color[marketplace_id=__MP__][language_tag=en_us]#1.value',
'part_number':'part_number[marketplace_id=__MP__]#1.value',
'fit_type':'fit_type[marketplace_id=__MP__][language_tag=en_us]#1.value',
'care_instructions':'care_instructions[marketplace_id=__MP__][language_tag=en_us]#1.value',
'closure_type':'closure[marketplace_id=__MP__]#1.type[language_tag=en_us]#1.value',
'size_name':'size[marketplace_id=__MP__][language_tag=en_us]#1.value',
'bottoms_size_system':'bottoms_size[marketplace_id=__MP__]#1.size_system',
'bottoms_size_class':'bottoms_size[marketplace_id=__MP__]#1.size_class',
'bottoms_size':'bottoms_size[marketplace_id=__MP__]#1.size',
'bottoms_size_to':'bottoms_size[marketplace_id=__MP__]#1.size_to',
'shirt_size_system':'shirt_size[marketplace_id=__MP__]#1.size_system',
'shirt_size_class':'shirt_size[marketplace_id=__MP__]#1.size_class',
'shirt_size':'shirt_size[marketplace_id=__MP__]#1.size',
'footwear_size_system':'footwear_size[marketplace_id=__MP__]#1.size_system',
'footwear_size_class':'footwear_size[marketplace_id=__MP__]#1.size_class',
'footwear_size':'footwear_size[marketplace_id=__MP__]#1.size',
'collar_style':'collar_style[marketplace_id=__MP__][language_tag=en_us]#1.value',
'leg_style':'leg[marketplace_id=__MP__]#1.style[language_tag=en_us]#1.value',
'rise_style':'rise[marketplace_id=__MP__]#1.style[language_tag=en_us]#1.value',
'outer_material_type':'outer[marketplace_id=__MP__]#1.material[language_tag=en_us]#1.value',
'inner_material_type':'inner[marketplace_id=__MP__]#1.material[language_tag=en_us]#1.value',
'main_image_url':'main_product_image_locator[marketplace_id=__MP__]#1.media_location',
'other_image_url1':'other_product_image_locator_1[marketplace_id=__MP__]#1.media_location',
'other_image_url2':'other_product_image_locator_2[marketplace_id=__MP__]#1.media_location',
'other_image_url3':'other_product_image_locator_3[marketplace_id=__MP__]#1.media_location',
'other_image_url4':'other_product_image_locator_4[marketplace_id=__MP__]#1.media_location',
'other_image_url5':'other_product_image_locator_5[marketplace_id=__MP__]#1.media_location',
'other_image_url6':'other_product_image_locator_6[marketplace_id=__MP__]#1.media_location',
'other_image_url7':'other_product_image_locator_7[marketplace_id=__MP__]#1.media_location',
'other_image_url8':'other_product_image_locator_8[marketplace_id=__MP__]#1.media_location',
'package_length_unit_of_measure':'item_package_dimensions[marketplace_id=__MP__]#1.length.unit',
'package_length':'item_package_dimensions[marketplace_id=__MP__]#1.length.value',
'package_width_unit_of_measure':'item_package_dimensions[marketplace_id=__MP__]#1.width.unit',
'package_width':'item_package_dimensions[marketplace_id=__MP__]#1.width.value',
'package_height_unit_of_measure':'item_package_dimensions[marketplace_id=__MP__]#1.height.unit',
'package_height':'item_package_dimensions[marketplace_id=__MP__]#1.height.value',
'package_weight_unit_of_measure':'item_package_weight[marketplace_id=__MP__]#1.unit',
'package_weight':'item_package_weight[marketplace_id=__MP__]#1.value',
'item_length':'item_dimensions[marketplace_id=__MP__]#1.length.value',
'item_length_unit_of_measure':'item_dimensions[marketplace_id=__MP__]#1.length.unit',
'item_width':'item_dimensions[marketplace_id=__MP__]#1.width.value',
'item_width_unit_of_measure':'item_dimensions[marketplace_id=__MP__]#1.width.unit',
'item_height':'item_dimensions[marketplace_id=__MP__]#1.height.value',
'item_height_unit_of_measure':'item_dimensions[marketplace_id=__MP__]#1.height.unit',
'item_weight':'item_weight[marketplace_id=__MP__]#1.value',
'item_weight_unit_of measure':'item_weight[marketplace_id=__MP__]#1.unit',
'target_audience_keywords':'target_audience_keyword[marketplace_id=__MP__][language_tag=en_us]#1.value',
'generic_keywords':'generic_keyword[marketplace_id=__MP__][language_tag=en_us]#1.value',
'subject_keyword':'subject_keyword[marketplace_id=__MP__][language_tag=en_us]#1.value',
'country_of_origin':'country_of_origin[marketplace_id=__MP__]#1.value',
'batteries_required':'batteries_required[marketplace_id=__MP__]#1.value',
'batteries_included':'batteries_included[marketplace_id=__MP__]#1.value',
'number_of_items':'number_of_items[marketplace_id=__MP__]#1.value',
'unit_count':'unit_count[marketplace_id=__MP__]#1.value',
'unit_count_type':'unit_count[marketplace_id=__MP__]#1.type[language_tag=en_us].value',
'quantity':'fulfillment_availability#1.quantity',
'fulfillment_availability':'fulfillment_availability#1.is_inventory_available',
'merchant_shipping_group_name':'merchant_shipping_group[marketplace_id=__MP__]#1.value',
'parent_child':'parentage_level[marketplace_id=__MP__]#1.value',
'relationship_type':'child_parent_sku_relationship[marketplace_id=__MP__]#1.child_relationship_type',
'parent_sku':'child_parent_sku_relationship[marketplace_id=__MP__]#1.parent_sku',
'variation_theme':'variation_theme#1.name',
'safety_warning':'safety_warning[marketplace_id=__MP__][language_tag=en_us]#1.value',
'warranty_description':'warranty_description[marketplace_id=__MP__][language_tag=en_us]#1.value',
'legal_disclaimer_description':'legal_disclaimer_description[marketplace_id=__MP__][language_tag=en_us]#1.value',
'is_expiration_dated_product':'is_expiration_dated_product[marketplace_id=__MP__]#1.value',
'product_site_launch_date':'product_site_launch_date[marketplace_id=__MP__]#1.value',
'offering_start_date':'purchasable_offer[marketplace_id=__MP__][audience=all]#1.start_at.value',
'min_price':'purchasable_offer[marketplace_id=__MP__][audience=all]#1.minimum_seller_allowed_price#1.schedule#1.value_with_tax',
'max_price':'purchasable_offer[marketplace_id=__MP__][audience=all]#1.maximum_seller_allowed_price#1.schedule#1.value_with_tax',
'specific_uses_for_product':'specific_uses_for_product[marketplace_id=__MP__][language_tag=en_us]#1.value',
'product_benefit':'product_benefit[marketplace_id=__MP__][language_tag=en_us]#1.value',
'short_product_description':'short_product_description[marketplace_id=__MP__][language_tag=en_us]#1.value',
'item_type':'item_type_keyword[marketplace_id=__MP__]#1.value',
'supplier_declared_dg_hz_regulation1':'supplier_declared_dg_hz_regulation[marketplace_id=__MP__]#1.value',
'special_features1':'special_feature[marketplace_id=__MP__][language_tag=en_us]#1.value',
'material_composition':'material_composition[marketplace_id=__MP__][language_tag=en_us]#1.value',
'included_components':'included_components[marketplace_id=__MP__][language_tag=en_us]#1.value',
'is_assembly_required':'is_assembly_required[marketplace_id=__MP__]#1.value',
'power_plug_type':'power_plug_type[marketplace_id=__MP__]#1.value',
'number_of_boxes':'number_of_boxes[marketplace_id=__MP__]#1.value',
'dsa_responsible_party_address':'dsa_responsible_party_address[marketplace_id=__MP__]#1.value',
'gpsr_safety_attestation':'gpsr_safety_attestation[marketplace_id=__MP__]#1.value',
}

MP_IDS = {'US':'atvpdkikx0der','CA':'a2euq1wtgctbg2','EU':'a1f83g8c2aro7p','AU':'a39ibj37trp1c6'}

def get_mapping(mp):
    mid = MP_IDS.get(mp, 'atvpdkikx0der')
    return {k: v.replace('__MP__', mid) for k, v in MAPPING_BASE.items()}

def cell_ref_to_row_col(ref):
    m = re.match(r'([A-Z]+)(\d+)', ref.upper())
    if m: return int(m.group(2)), column_index_from_string(m.group(1))
    return None, None

def xml_escape(v):
    return str(v).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')

def get_cell_style_idx(sheet_xml, row, col):
    ref = f"{get_column_letter(col)}{row}"
    m = re.search(rf'<c r="{re.escape(ref)}"([^>]*)>', sheet_xml)
    if not m: return None
    s = re.search(r'\bs="(\d+)"', m.group(1))
    return int(s.group(1)) if s else None

def get_fill_for_style(styles_xml, style_idx):
    xf_section = re.search(r'<cellXfs[^>]*>(.*?)</cellXfs>', styles_xml, re.DOTALL)
    if not xf_section: return None
    xfs = re.findall(r'<xf\b.*?(?:/>|</xf>)', xf_section.group(1), re.DOTALL)
    if style_idx >= len(xfs): return None
    fill_id_m = re.search(r'fillId="(\d+)"', xfs[style_idx])
    if not fill_id_m: return None
    fill_id = int(fill_id_m.group(1))
    fills = re.findall(r'<fill>.*?</fill>', styles_xml, re.DOTALL)
    if fill_id >= len(fills): return None
    fill = fills[fill_id]
    if 'patternType="none"' in fill or 'gray125' in fill or 'fgColor' not in fill:
        return None
    return fill

def ensure_fill_in_beta(beta_styles, fill_xml, cache):
    key = re.sub(r'\s+', ' ', fill_xml.strip())
    if key in cache: return cache[key], beta_styles
    fc_m = re.search(r'<fills count="(\d+)">', beta_styles)
    if not fc_m: return None, beta_styles
    fc = int(fc_m.group(1))
    beta_styles = beta_styles.replace(f'<fills count="{fc}">', f'<fills count="{fc+1}">',1)
    beta_styles = beta_styles.replace('</fills>', fill_xml+'</fills>',1)
    xc_m = re.search(r'<cellXfs count="(\d+)">', beta_styles)
    if not xc_m: return None, beta_styles
    xc = int(xc_m.group(1))
    beta_styles = beta_styles.replace(f'<cellXfs count="{xc}">', f'<cellXfs count="{xc+1}">',1)
    beta_styles = beta_styles.replace('</cellXfs>', f'<xf numFmtId="0" fontId="0" fillId="{fc}" borderId="0" xfId="0" applyFill="1"/></cellXfs>',1)
    cache[key] = xc
    return xc, beta_styles

def build_summary_sheet_xml(mapping_results, unmapped_std_cols, std_rows, std_col_name_by_idx):
    """Build enhanced 3-section summary worksheet XML.
    
    Section 1: Auto-transferred columns (mapped + found in Beta)
    Section 2: In mapping but Beta column not found (needs investigation)
    Section 3: Not in mapping at all — Standard columns with data needing manual transfer
    """
    rows_xml = ''
    row_cursor = [1]  # mutable so helper can increment

    def tc(col, row, val):
        return f'<c r="{get_column_letter(col)}{row}" t="inlineStr"><is><t>{xml_escape(str(val))}</t></is></c>'
    def nc(col, row, val):
        return f'<c r="{get_column_letter(col)}{row}"><v>{val}</v></c>'
    def section_header(label):
        r = row_cursor[0]
        s = f'<row r="{r}"><c r="A{r}" t="inlineStr"><is><t>{label}</t></is></c></row>'
        row_cursor[0] += 1
        return s
    def blank():
        r = row_cursor[0]; row_cursor[0] += 1
        return f'<row r="{r}"></row>'
    def col_header(cols):
        r = row_cursor[0]
        s = f'<row r="{r}">' + ''.join(tc(i+1, r, h) for i,h in enumerate(cols)) + '</row>'
        row_cursor[0] += 1
        return s
    def data_row(cells_dict):
        r = row_cursor[0]
        s = f'<row r="{r}">' + ''.join(v for v in cells_dict) + '</row>'
        row_cursor[0] += 1
        return s

    # ── TITLE ──
    rows_xml += section_header('⚡ NEXUS — Transfer Summary')
    rows_xml += f'<row r="{row_cursor[0]}"><c r="A{row_cursor[0]}" t="inlineStr"><is><t>Generated: {datetime.now().strftime("%d %b %Y %H:%M")} | Data rows: {len(std_rows)}</t></is></c></row>'
    row_cursor[0] += 1
    rows_xml += blank()

    # ── SECTION 1: MAPPED ──
    mapped = sorted([r for r in mapping_results if r['status'] == 'mapped'], key=lambda x: x['std_col'])
    rows_xml += section_header(f'✅  SECTION 1 — AUTO-TRANSFERRED  ({len(mapped)} columns)  — Data + Colors + Comments all moved')
    rows_xml += col_header(['Std Col', 'Standard Header', 'Beta Col', 'Beta Column Name', 'Cells Moved'])
    for r in mapped:
        rows_xml += data_row([
            tc(1, row_cursor[0], get_column_letter(r['std_col'])),
            tc(2, row_cursor[0], r['std_name']),
            tc(3, row_cursor[0], get_column_letter(r['beta_col'])),
            tc(4, row_cursor[0], r['beta_name']),
            nc(5, row_cursor[0], r['cells']),
        ])
    rows_xml += blank()

    # ── SECTION 2: NOT FOUND IN BETA ──
    not_found = sorted([r for r in mapping_results if r['status'] == 'not_found'], key=lambda x: x['std_col'])
    rows_xml += section_header(f'❌  SECTION 2 — IN MAPPING BUT BETA COLUMN NOT FOUND  ({len(not_found)})  — Check if Beta template version matches')
    if not_found:
        rows_xml += col_header(['Std Col', 'Standard Header', 'Expected Beta Column Name', 'Action'])
        for r in not_found:
            rows_xml += data_row([
                tc(1, row_cursor[0], get_column_letter(r['std_col'])),
                tc(2, row_cursor[0], r['std_name']),
                tc(3, row_cursor[0], r['beta_name']),
                tc(4, row_cursor[0], 'Beta column name may differ — verify header row 5'),
            ])
    else:
        r = row_cursor[0]; row_cursor[0] += 1
        rows_xml += f'<row r="{r}"><c r="A{r}" t="inlineStr"><is><t>None — all mapping entries found in Beta!</t></is></c></row>'
    rows_xml += blank()

    # ── SECTION 3: NOT MAPPED — only cols with actual data ──
    cols_with_data = []
    for col_idx, header in sorted(unmapped_std_cols, key=lambda x: x[0]):
        vals = [str(row[col_idx-1]) for row in std_rows
                if col_idx-1 < len(row) and row[col_idx-1] is not None and str(row[col_idx-1]).strip()]
        if vals:
            cols_with_data.append((col_idx, header, vals))

    rows_xml += section_header(f'📋  SECTION 3 — NOT IN MAPPING  ({len(cols_with_data)} columns with data)  — Manual transfer or update mapping code')
    if cols_with_data:
        rows_xml += col_header(['Std Col', 'Standard Header', 'Sample Value', 'Cells With Data', 'Action'])
        for col_idx, header, vals in cols_with_data:
            rows_xml += data_row([
                tc(1, row_cursor[0], get_column_letter(col_idx)),
                tc(2, row_cursor[0], header),
                tc(3, row_cursor[0], vals[0][:80]),
                nc(4, row_cursor[0], len(vals)),
                tc(5, row_cursor[0], 'Add to MAPPING_BASE in code or copy manually'),
            ])
    rows_xml += blank()

    # ── TOTALS ──
    total_cells = sum(r['cells'] for r in mapped)
    rows_xml += data_row([
        tc(1, row_cursor[0], 'TOTAL'),
        tc(2, row_cursor[0],
           f'✅ {len(mapped)} cols auto-transferred ({total_cells} cells)  |  '
           f'❌ {len(not_found)} cols not found in Beta  |  '
           f'📋 {len(cols_with_data)} cols need manual work'),
    ])

    ws_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<sheetViews><sheetView tabSelected="0" workbookViewId="0"/></sheetViews>
<sheetFormatPr defaultRowHeight="15"/>
<cols>
<col min="1" max="1" width="10" customWidth="1"/>
<col min="2" max="2" width="34" customWidth="1"/>
<col min="3" max="3" width="10" customWidth="1"/>
<col min="4" max="4" width="60" customWidth="1"/>
<col min="5" max="5" width="15" customWidth="1"/>
</cols>
<sheetData>{rows_xml}</sheetData>
</worksheet>'''
    return ws_xml, len(mapped), len(not_found), len(cols_with_data)

def convert(std_bytes, orig_beta_bytes, marketplace, std_hdr, std_data, beta_hdr, beta_data):
    mapping = get_mapping(marketplace)
    TC_NS = 'http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments'

    # Read Standard
    std_wb = load_workbook(io.BytesIO(std_bytes), read_only=True)
    std_ws = std_wb['Template'] if 'Template' in std_wb.sheetnames else std_wb.active
    std_rows = [list(r) for r in std_ws.iter_rows(min_row=std_data, values_only=True) if any(v is not None for v in r)]
    std_col_map_raw = {str(c.value).strip().lower(): c.column for c in std_ws[std_hdr] if c.value}
    std_col_name_by_idx = {c.column: str(c.value).strip() for c in std_ws[std_hdr] if c.value}
    std_wb.close()

    with zipfile.ZipFile(io.BytesIO(std_bytes), 'r') as z:
        files = z.namelist()
        tc_xml    = z.read(next(f for f in files if 'threadedcomment' in f.lower())).decode('utf-8')
        pers_xml  = z.read(next((f for f in files if 'person' in f.lower()), None) or '').decode('utf-8') if any('person' in f.lower() for f in files) else ''
        std_styles = z.read('xl/styles.xml').decode('utf-8')
        std_wb_xml  = z.read('xl/workbook.xml').decode('utf-8')
        std_wb_rels = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        rid = re.search(r'<sheet[^>]*name="Template"[^>]*r:id="([^"]+)"', std_wb_xml)
        if rid:
            rid = rid.group(1)
            for m2 in re.finditer(r'<Relationship([^>]*)>', std_wb_rels):
                if rid in m2.group(1):
                    t = re.search(r'Target="([^"]+)"', m2.group(1))
                    if t: std_sheet_xml = z.read(f'xl/{t.group(1)}').decode('utf-8'); break
        else:
            std_sheet_xml = ''

    # Find Beta info
    with zipfile.ZipFile(io.BytesIO(orig_beta_bytes), 'r') as z:
        wb_xml    = z.read('xl/workbook.xml').decode('utf-8')
        wb_rels   = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        all_orig  = z.namelist()
        beta_styles_xml = z.read('xl/styles.xml').decode('utf-8')
    
    beta_wb = load_workbook(io.BytesIO(orig_beta_bytes), read_only=True)
    beta_ws = next((beta_wb[n] for n in beta_wb.sheetnames if n.lower()=='template'), beta_wb.active)
    beta_col_map_raw = {str(c.value).strip().lower(): c.column for c in beta_ws[beta_hdr] if c.value}
    beta_col_name_by_idx = {c.column: str(c.value).strip() for c in beta_ws[beta_hdr] if c.value}
    beta_wb.close()

    tmpl_rid = re.search(r'<sheet[^>]*name="Template"[^>]*r:id="([^"]+)"', wb_xml).group(1)
    for m in re.finditer(r'<Relationship([^>]*)>', wb_rels):
        if tmpl_rid in m.group(1):
            t = re.search(r'Target="([^"]+)"', m.group(1))
            if t: sheet_target = t.group(1); break
    sheet_num = re.search(r'sheet(\d+)\.xml', sheet_target).group(1)

    TC_PATH   = 'xl/threadedComments/threadedComment1.xml'
    LEG_PATH  = f'xl/comments{sheet_num}.xml'
    VML_PATH  = f'xl/drawings/vmlDrawingTC{sheet_num}.vml'
    PERS_PATH = 'xl/persons/person.xml'
    RELS_PATH = f'xl/worksheets/_rels/sheet{sheet_num}.xml.rels'
    WS_PATH   = f'xl/worksheets/sheet{sheet_num}.xml'
    VML_RID   = 'rIdVML_TC'

    # Build col pairs + mapping results for summary
    col_pairs = []
    mapping_results = []
    for std_name, beta_full in mapping.items():
        std_c = std_col_map_raw.get(std_name.lower())
        beta_c = beta_col_map_raw.get(beta_full.lower())
        status = 'mapped' if (std_c and beta_c) else 'not_found'
        if std_c:
            cells_count = sum(1 for row in std_rows if std_c-1 < len(row) and row[std_c-1] is not None and str(row[std_c-1]).strip())
            mapping_results.append({'std_col': std_c,'std_name': std_col_name_by_idx.get(std_c, std_name),'beta_col': beta_c,'beta_name': beta_full,'status': status,'cells': cells_count if status=='mapped' else 0})
            if std_c and beta_c:
                col_pairs.append((std_c, beta_c))

    # Remap comments
    import xml.etree.ElementTree as ET
    root = ET.fromstring(tc_xml)
    col_map = {s: b for s, b in col_pairs}
    remapped = []
    for tc in root:
        ref = tc.get('ref','')
        sr, sc = cell_ref_to_row_col(ref)
        if sr is None or sr < std_data: continue
        bc = col_map.get(sc)
        if not bc: continue
        br = beta_data + (sr - std_data)
        beta_ref = f"{get_column_letter(bc)}{br}"
        txt_el = tc.find(f'{{{TC_NS}}}text')
        remapped.append({'ref': beta_ref,'id': tc.get('id',''),'pid': tc.get('personId',''),'dT': tc.get('dT',''),'txt': txt_el.text if txt_el is not None else ''})

    # Build color map
    fill_cache = {}
    beta_cell_styles = {}
    for ridx, std_row in enumerate(std_rows):
        for sc, bc in col_pairs:
            si = get_cell_style_idx(std_sheet_xml, std_data + ridx, sc)
            if si is None: continue
            fill_xml = get_fill_for_style(std_styles, si)
            if not fill_xml: continue
            new_si, beta_styles_xml = ensure_fill_in_beta(beta_styles_xml, fill_xml, fill_cache)
            if new_si is not None:
                beta_cell_styles[(ridx, sc)] = new_si

    # Build comment XMLs
    tc_lines = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n<ThreadedComments xmlns="http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments" xmlns:x="http://schemas.openxmlformats.org/spreadsheetml/2006/main">']
    for c in remapped:
        t = xml_escape(c['txt'])
        tc_lines.append(f'<threadedComment ref="{c["ref"]}" dT="{c["dT"]}" personId="{c["pid"]}" id="{c["id"]}"><text xml:space="preserve">{t}</text></threadedComment>')
    tc_lines.append('</ThreadedComments>')
    new_tc_xml = '\r\n'.join(tc_lines)

    id_list, id_idx = [], {}
    for c in remapped:
        k = c['id'].strip('{}')
        if k not in id_idx: id_idx[k] = len(id_list); id_list.append(k)
    leg = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n<comments xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><authors>']
    for tid in id_list: leg.append(f'<author>tc={{{tid}}}</author>')
    leg.append('</authors><commentList>')
    for c in remapped:
        k = c['id'].strip('{}'); t = xml_escape(c['txt'])
        leg.append(f'<comment ref="{c["ref"]}" authorId="{id_idx[k]}" shapeId="0"><text><r><t xml:space="preserve">{t}</t></r></text></comment>')
    leg.append('</commentList></comments>')
    new_leg_xml = '\r\n'.join(leg)

    vml = ['<xml xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office" xmlns:x="urn:schemas-microsoft-com:office:excel"><o:shapelayout v:ext="edit"><o:idmap v:ext="edit" data="1"/></o:shapelayout><v:shapetype id="_x0000_t202" coordsize="21600,21600" o:spt="202" path="m,l,21600r21600,l21600,xe"><v:stroke joinstyle="miter"/><v:path gradientshapeok="t" o:connecttype="rect"/></v:shapetype>']
    for i, c in enumerate(remapped):
        r, col = cell_ref_to_row_col(c['ref'])
        if r is None: continue
        vml.append(f'<v:shape id="_x0000_s{1025+i}" type="#_x0000_t202" style=\'position:absolute;margin-left:59.25pt;margin-top:1.5pt;width:108pt;height:59.25pt;z-index:{i+1};visibility:hidden\' fillcolor="#ffffe1" o:insetmode="auto"><v:fill color2="#ffffe1"/><v:shadow color="black" obscured="t"/><v:path o:connecttype="none"/><v:textbox style=\'mso-direction-alt:auto\'><div style=\'text-align:left\'></div></v:textbox><x:ClientData ObjectType="Note"><x:MoveWithCells/><x:SizeWithCells/><x:Anchor>{col},0,{r-1},0,{col+2},5,{r+5},5</x:Anchor><x:AutoFill>False</x:AutoFill><x:Row>{r-1}</x:Row><x:Column>{col-1}</x:Column></x:ClientData></v:shape>')
    vml.append('</xml>')
    new_vml_xml = ''.join(vml)

    # Build summary sheet — with unmapped columns
    mapped_std_col_set = {r['std_col'] for r in mapping_results}
    unmapped_std_cols = [(ci, n) for ci, n in std_col_name_by_idx.items() if ci not in mapped_std_col_set]
    summary_xml, mapped_count, not_found_count, manual_count = build_summary_sheet_xml(
        mapping_results, unmapped_std_cols, std_rows, std_col_name_by_idx)

    # Build output zip
    out_buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(orig_beta_bytes), 'r') as zin:
        with zipfile.ZipFile(out_buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            # Find new sheet number for summary
            sheet_nums = [int(m) for f in all_orig for m in re.findall(r'sheet(\d+)\.xml', f)]
            summary_sheet_num = max(sheet_nums, default=10) + 1
            summary_path = f'xl/worksheets/sheet{summary_sheet_num}.xml'
            SUMMARY_RID = f'rIdSummary'

            for item in all_orig:
                data = zin.read(item)

                if item == WS_PATH:
                    ws = data.decode('utf-8')
                    new_rows = ''
                    for ridx, std_row in enumerate(std_rows):
                        br = beta_data + ridx
                        col_vals = {}
                        for sc, bc in col_pairs:
                            val = std_row[sc-1] if sc-1 < len(std_row) else None
                            if val is not None and str(val).strip():
                                col_vals[bc] = (val, beta_cell_styles.get((ridx, sc)))
                        if col_vals:
                            cells = ''.join(f'<c r="{get_column_letter(bc)}{br}"{f" s=\"{si}\"" if si else ""} t="inlineStr"><is><t>{xml_escape(str(v))}</t></is></c>' for bc,(v,si) in sorted(col_vals.items()))
                            new_rows += f'<row r="{br}">{cells}</row>'
                    sd_s = ws.find('<sheetData>'); sd_e = ws.find('</sheetData>')
                    if sd_s >= 0 and sd_e >= 0:
                        kept = ''.join(m.group(0) for m in re.finditer(r'<row r="(\d+)"[^>]*>.*?</row>', ws[sd_s+len('<sheetData>'):sd_e], re.DOTALL) if int(m.group(1)) < beta_data)
                        ws = ws[:sd_s] + f'<sheetData>{kept}{new_rows}</sheetData>' + ws[sd_e+len('</sheetData>'):]
                    if 'legacyDrawing' not in ws:
                        target = '<extLst>' if '<extLst>' in ws else '</worksheet>'
                        ws = ws.replace(target, f'<legacyDrawing r:id="{VML_RID}"/>{target}', 1)
                    zout.writestr(item, ws)

                elif item == 'xl/styles.xml':
                    zout.writestr(item, beta_styles_xml)

                elif item == 'xl/workbook.xml':
                    wb = data.decode('utf-8')
                    # Add summary sheet to workbook
                    wb = wb.replace('</sheets>', f'<sheet name="Transfer Summary" sheetId="{summary_sheet_num}" r:id="{SUMMARY_RID}"/></sheets>')
                    zout.writestr(item, wb)

                elif item == 'xl/_rels/workbook.xml.rels':
                    rels = data.decode('utf-8')
                    if 'person' not in rels.lower() and pers_xml:
                        nid = max([int(x) for x in re.findall(r'rId(\d+)', rels) if x.isdigit()], default=20) + 1
                        rels = rels.replace('</Relationships>', f'<Relationship Id="rId{nid}" Type="http://schemas.microsoft.com/office/2017/10/relationships/person" Target="persons/person.xml"/></Relationships>')
                    # Add summary sheet rel
                    rels = rels.replace('</Relationships>', f'<Relationship Id="{SUMMARY_RID}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{summary_sheet_num}.xml"/></Relationships>')
                    zout.writestr(item, rels)

                elif item == RELS_PATH:
                    rels = data.decode('utf-8')
                    existing = [int(x) for x in re.findall(r'rId(\d+)', rels) if x.isdigit()]
                    nid = max(existing, default=0) + 1
                    nr = ''
                    if 'threadedComment' not in rels: nr += f'<Relationship Id="rId{nid}" Type="http://schemas.microsoft.com/office/2017/10/relationships/threadedComment" Target="../threadedComments/threadedComment1.xml"/>'; nid += 1
                    if 'relationships/comments' not in rels: nr += f'<Relationship Id="rId{nid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/comments" Target="../{LEG_PATH.replace("xl/","")}"/>'; nid += 1
                    if VML_RID not in rels: nr += f'<Relationship Id="{VML_RID}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing" Target="../drawings/vmlDrawingTC{sheet_num}.vml"/>'
                    if nr: rels = rels.replace('</Relationships>', nr + '</Relationships>')
                    zout.writestr(item, rels)

                elif item == '[Content_Types].xml':
                    ct = data.decode('utf-8')
                    adds = ''
                    if 'threadedcomments' not in ct.lower(): adds += f'<Override PartName="/{TC_PATH}" ContentType="application/vnd.ms-excel.threadedcomments+xml"/>'
                    if 'person' not in ct.lower() and pers_xml: adds += f'<Override PartName="/{PERS_PATH}" ContentType="application/vnd.ms-excel.person+xml"/>'
                    adds += f'<Override PartName="/{LEG_PATH}" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.comments+xml"/>'
                    adds += f'<Override PartName="/{VML_PATH}" ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing"/>'
                    adds += f'<Override PartName="/{summary_path}" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
                    ct = ct.replace('</Types>', adds + '</Types>')
                    zout.writestr(item, ct)

                else:
                    zout.writestr(item, data)

            # Write new files
            zout.writestr(TC_PATH, new_tc_xml)
            zout.writestr(LEG_PATH, new_leg_xml)
            zout.writestr(VML_PATH, new_vml_xml)
            zout.writestr(summary_path, summary_xml)
            if pers_xml: zout.writestr(PERS_PATH, pers_xml)
            if RELS_PATH not in all_orig:
                zout.writestr(RELS_PATH, f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.microsoft.com/office/2017/10/relationships/threadedComment" Target="../threadedComments/threadedComment1.xml"/><Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/comments" Target="../{LEG_PATH.replace("xl/","")}"/><Relationship Id="{VML_RID}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing" Target="../drawings/vmlDrawingTC{sheet_num}.vml"/></Relationships>')

    out_buf.seek(0)
    return out_buf, len(std_rows), len(col_pairs), len(remapped), len(fill_cache), mapped_count, not_found_count, manual_count


# ── HTML UI (same design as before) ──
HTML = open('/mnt/user-data/outputs/NEXUS_Standard_to_Beta_Converter.html').read() if False else '''<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>NEXUS — Standard to Beta Converter v2.0</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=DM+Sans:wght@300;400;500&family=DM+Mono:wght@400;500&display=swap');
:root{--navy:#0D1B2A;--gold:#F4B942;--gold-light:#F8D07A;--blue:#2E75B6;--blue-light:#5BA3E0;--white:#F0F4F8;--gray:#8BA0B4;--success:#4CAF82;--card:rgba(26,46,69,0.9);}
*{margin:0;padding:0;box-sizing:border-box;}
body{background:var(--navy);color:var(--white);font-family:"DM Sans",sans-serif;min-height:100vh;}
body::before{content:"";position:fixed;inset:0;background-image:linear-gradient(rgba(46,117,182,0.04) 1px,transparent 1px),linear-gradient(90deg,rgba(46,117,182,0.04) 1px,transparent 1px);background-size:40px 40px;pointer-events:none;}
.orb{position:fixed;border-radius:50%;filter:blur(80px);pointer-events:none;}
.o1{width:500px;height:500px;background:rgba(46,117,182,0.12);top:-100px;left:-100px;}
.o2{width:400px;height:400px;background:rgba(244,185,66,0.06);bottom:-80px;right:-80px;}
.wrap{position:relative;z-index:1;max-width:820px;margin:0 auto;padding:40px 24px 80px;}
header{text-align:center;margin-bottom:44px;}
.badge{display:inline-flex;align-items:center;gap:12px;margin-bottom:20px;}
.logo{background:linear-gradient(135deg,var(--gold),var(--gold-light));color:var(--navy);font-family:"Syne",sans-serif;font-weight:800;font-size:12px;letter-spacing:.12em;padding:6px 13px;border-radius:6px;}
.sep{width:1px;height:22px;background:rgba(240,244,248,.15);}
.sub{font-family:"DM Mono",monospace;font-size:11px;color:var(--gray);letter-spacing:.08em;}
h1{font-family:"Syne",sans-serif;font-weight:800;font-size:clamp(24px,5vw,38px);background:linear-gradient(135deg,var(--white),var(--blue-light));-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;margin-bottom:10px;}
.desc{color:var(--gray);font-size:14px;max-width:520px;margin:0 auto;line-height:1.6;}
.lbl{font-family:"DM Mono",monospace;font-size:10px;letter-spacing:.14em;text-transform:uppercase;color:var(--gold);margin-bottom:10px;}
.sec{margin-bottom:24px;}
.g4{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;}
.g2{display:grid;grid-template-columns:1fr 1fr;gap:14px;}
.card{background:var(--card);border:1.5px solid rgba(240,244,248,.08);border-radius:12px;padding:16px;backdrop-filter:blur(8px);}
.mp{cursor:pointer;text-align:center;border-radius:10px;padding:13px 8px;transition:.2s;}
.mp:hover{border-color:rgba(46,117,182,.4);}
.mp.on{border-color:var(--gold);background:rgba(244,185,66,.08);}
.flag{font-size:20px;margin-bottom:3px;}
.mn{font-family:"Syne",sans-serif;font-weight:700;font-size:12px;}
.mid{font-family:"DM Mono",monospace;font-size:9px;color:var(--gray);margin-top:2px;}
.mp.on .mn{color:var(--gold);}
.rl{font-family:"DM Mono",monospace;font-size:10px;color:var(--gold);letter-spacing:.1em;text-transform:uppercase;margin-bottom:7px;}
.ri{width:100%;background:rgba(13,27,42,.8);border:1px solid rgba(240,244,248,.12);border-radius:6px;color:var(--white);font-family:"DM Mono",monospace;font-size:20px;font-weight:700;padding:8px;text-align:center;outline:none;}
.rh{font-size:11px;color:var(--gray);margin-top:4px;text-align:center;}
.up{cursor:pointer;text-align:center;padding:24px 16px;transition:.25s;}
.up:hover{border-color:rgba(46,117,182,.4);transform:translateY(-2px);}
.up.ok{border-color:rgba(76,175,130,.5);background:rgba(76,175,130,.05);}
.ico{width:48px;height:48px;border-radius:11px;display:flex;align-items:center;justify-content:center;margin:0 auto 12px;font-size:20px;background:rgba(46,117,182,.15);border:1px solid rgba(46,117,182,.2);}
.ok .ico{background:rgba(76,175,130,.15);border-color:rgba(76,175,130,.2);}
.ut{font-family:"Syne",sans-serif;font-weight:700;font-size:14px;margin-bottom:5px;}
.ud{font-size:12px;color:var(--gray);line-height:1.5;margin-bottom:12px;}
.utag{display:inline-block;font-family:"DM Mono",monospace;font-size:10px;padding:3px 10px;border-radius:20px;background:rgba(46,117,182,.15);color:var(--blue-light);border:1px solid rgba(46,117,182,.2);}
.ok .utag{background:rgba(76,175,130,.15);color:var(--success);border-color:rgba(76,175,130,.2);}
.fn{font-family:"DM Mono",monospace;font-size:10px;color:var(--success);margin-top:6px;word-break:break-all;}
input[type=file]{display:none;}
.btn{width:100%;border:none;border-radius:12px;padding:17px;font-family:"Syne",sans-serif;font-weight:800;font-size:15px;letter-spacing:.06em;text-transform:uppercase;cursor:pointer;transition:.25s;}
.bg{background:linear-gradient(135deg,var(--gold),var(--gold-light));color:var(--navy);}
.bg:hover{transform:translateY(-2px);box-shadow:0 8px 28px rgba(244,185,66,.3);}
.bg:disabled{opacity:.4;cursor:not-allowed;transform:none;box-shadow:none;}
.bv{background:linear-gradient(135deg,var(--success),#6BCFA0);color:var(--navy);display:none;margin-top:14px;}
.bv:hover{transform:translateY(-2px);box-shadow:0 8px 28px rgba(76,175,130,.3);}
.pg{display:none;margin:16px 0;}
.pgb{background:rgba(240,244,248,.06);border-radius:100px;height:5px;overflow:hidden;margin-bottom:7px;}
.pf{height:100%;background:linear-gradient(90deg,var(--blue),var(--gold));border-radius:100px;width:0%;transition:width .4s;}
.pt{font-family:"DM Mono",monospace;font-size:11px;color:var(--gray);text-align:center;}
.log{display:none;background:rgba(13,27,42,.9);border:1px solid rgba(240,244,248,.06);border-radius:10px;padding:16px;margin:14px 0;max-height:200px;overflow-y:auto;font-family:"DM Mono",monospace;font-size:11px;line-height:1.9;}
.stats{display:none;grid-template-columns:repeat(4,1fr);gap:10px;margin:14px 0;}
.stat{background:var(--card);border:1px solid rgba(240,244,248,.06);border-radius:10px;padding:14px;text-align:center;}
.sv{font-family:"Syne",sans-serif;font-weight:800;font-size:22px;margin-bottom:3px;}
.sl{font-family:"DM Mono",monospace;font-size:9px;color:var(--gray);letter-spacing:.07em;text-transform:uppercase;line-height:1.4;}
.blue{color:var(--blue-light);}.gold{color:var(--gold);}.green{color:var(--success);}.orange{color:#F4A742;}.red{color:#E05252;}
.features{display:flex;gap:8px;justify-content:center;flex-wrap:wrap;margin-top:14px;}
.feat{font-family:"DM Mono",monospace;font-size:10px;padding:4px 10px;border-radius:20px;background:rgba(46,117,182,.12);color:var(--blue-light);border:1px solid rgba(46,117,182,.15);}
footer{text-align:center;margin-top:56px;padding-top:20px;border-top:1px solid rgba(240,244,248,.06);font-family:"DM Mono",monospace;font-size:10px;color:rgba(139,160,180,.4);letter-spacing:.07em;}
footer span{color:var(--gold);}
@media(max-width:580px){.g4,.g2,.stats{grid-template-columns:1fr 1fr;}}
</style>
</head>
<body>
<div class="orb o1"></div><div class="orb o2"></div>
<div class="wrap">
<header>
  <div class="badge"><div class="logo">⚡ NEXUS</div><div class="sep"></div><div class="sub">v2.0 · Ops-Avengers · Pattern</div></div>
  <h1>Standard → Beta Converter</h1>
  <p class="desc">Transfers data, threaded comments, cell colors, and generates a summary tab of all mapped vs unmapped columns.</p>
  <div class="features">
    <span class="feat">📋 Data Transfer</span>
    <span class="feat">💬 Comments</span>
    <span class="feat">🎨 Colors</span>
    <span class="feat">📊 Summary Tab</span>
    <span class="feat">🌍 US · CA · EU · AU</span>
  </div>
</header>

<form id="form" action="/convert" method="post" enctype="multipart/form-data">
<div class="sec">
  <div class="lbl">01 — Marketplace</div>
  <div class="g4">
    <div class="card mp on" data-mp="US" onclick="selMP(this)"><div class="flag">🇺🇸</div><div class="mn">US</div><div class="mid">ATVPDKIKX0DER</div></div>
    <div class="card mp" data-mp="CA" onclick="selMP(this)"><div class="flag">🇨🇦</div><div class="mn">CA</div><div class="mid">A2EUQ1WTGCTBG2</div></div>
    <div class="card mp" data-mp="EU" onclick="selMP(this)"><div class="flag">🇪🇺</div><div class="mn">EU</div><div class="mid">A1F83G8C2ARO7P</div></div>
    <div class="card mp" data-mp="AU" onclick="selMP(this)"><div class="flag">🇦🇺</div><div class="mn">AU</div><div class="mid">A39IBJ37TRP1C6</div></div>
  </div>
  <input type="hidden" name="marketplace" id="mp_input" value="US">
</div>

<div class="sec">
  <div class="lbl">02 — Header Rows</div>
  <div class="g4">
    <div class="card"><div class="rl">Std Header Row</div><input class="ri" type="number" name="std_hdr" value="3" min="1" max="20"><div class="rh">Standard = Row 3</div></div>
    <div class="card"><div class="rl">Std Data Starts</div><input class="ri" type="number" name="std_data" value="4" min="1" max="20"><div class="rh">Standard = Row 4</div></div>
    <div class="card"><div class="rl">Beta Header Row</div><input class="ri" type="number" name="beta_hdr" value="5" min="1" max="20"><div class="rh">Beta English = Row 5</div></div>
    <div class="card"><div class="rl">Beta Data Starts</div><input class="ri" type="number" name="beta_data" value="7" min="1" max="20"><div class="rh">Beta = Row 7</div></div>
  </div>
</div>

<div class="sec">
  <div class="lbl">03 — Upload Files</div>
  <div class="g2">
    <label class="card up" id="cu" for="fu">
      <div class="ico">📋</div><div class="ut">Standard Template</div>
      <div class="ud">Your filled Standard file with all data, colors and QC comments</div>
      <div class="utag" id="tu">Click to browse</div><div class="fn" id="nu"></div>
    </label>
    <input type="file" id="fu" name="standard" accept=".xlsx,.xls,.xlsm" onchange="setFile(this,'u')">
    <label class="card up" id="cb" for="fb">
      <div class="ico">🚀</div><div class="ut">Beta Template</div>
      <div class="ud">Fresh empty Beta template downloaded from Amazon</div>
      <div class="utag" id="tb">Click to browse</div><div class="fn" id="nb"></div>
    </label>
    <input type="file" id="fb" name="beta" accept=".xlsx,.xls,.xlsm" onchange="setFile(this,'b')">
  </div>
</div>

<div class="sec">
  <div class="lbl">04 — Convert</div>
  <button type="submit" class="btn bg" id="convBtn" disabled>⚡ Convert Standard → Beta</button>
  <div class="pg" id="pg"><div class="pgb"><div class="pf" id="pf"></div></div><div class="pt" id="pt">Processing...</div></div>
  <div class="log" id="log"></div>
  <div class="stats" id="stats">
    <div class="stat"><div class="sv blue" id="s1">0</div><div class="sl">Rows</div></div>
    <div class="stat"><div class="sv gold" id="s2">0</div><div class="sl">Cols Mapped</div></div>
    <div class="stat"><div class="sv green" id="s3">0</div><div class="sl">Comments</div></div>
    <div class="stat"><div class="sv orange" id="s4">0</div><div class="sl">Need Review</div></div>
  </div>
  <button type="button" class="btn bv" id="dlBtn">⬇ Download Converted Beta File</button>
</div>
</form>

<footer><p>Built by <span>Karan Singh</span> · <span>NEXUS v2.0</span> · Ops-Avengers · Pattern</p></footer>
</div>

<script>
let ok={u:false,b:false};
function selMP(el){document.querySelectorAll('.mp').forEach(b=>b.classList.remove('on'));el.classList.add('on');document.getElementById('mp_input').value=el.dataset.mp;}
function setFile(inp,t){const f=inp.files[0];if(!f)return;document.getElementById('c'+t).classList.add('ok');document.getElementById('t'+t).textContent='✓ File loaded';document.getElementById('n'+t).textContent=f.name;ok[t]=true;if(ok.u&&ok.b)document.getElementById('convBtn').disabled=false;}
document.getElementById('form').onsubmit=async function(e){
  e.preventDefault();
  const btn=document.getElementById('convBtn');btn.disabled=true;
  const pg=document.getElementById('pg'),log=document.getElementById('log');
  pg.style.display='block';log.style.display='block';log.innerHTML='';
  document.getElementById('stats').style.display='none';document.getElementById('dlBtn').style.display='none';
  function setP(p,t){document.getElementById('pf').style.width=p+'%';document.getElementById('pt').textContent=t;}
  function addL(m,c=''){log.innerHTML+=`<span style="color:${c||'var(--gray)'}">▸ ${m}</span><br>`;log.scrollTop=log.scrollHeight;}
  setP(20,'Uploading...');addL('Uploading files...');
  const fd=new FormData(this);
  try{
    setP(50,'Converting data, colors and comments...');addL('Converting...');
    const res=await fetch('/convert',{method:'POST',body:fd});
    if(!res.ok){const err=await res.json();addL('ERROR: '+err.error,'var(--gold)');setP(0,'Error!');btn.disabled=false;return;}
    setP(90,'Finalizing...');
    const info=JSON.parse(res.headers.get('X-Stats')||'{}');
    addL(`Rows: ${info.rows}`,'var(--success)');
    addL(`Columns mapped: ${info.mapped}`,'var(--success)');
    addL(`Comments: ${info.comments}`,'#B47FE8');
    addL(`Colors: ${info.colors} unique styles`,'#F4A742');
    addL(`Not found in Beta: ${info.not_found} → see Section 2 of Summary tab`,'var(--gray)');
    addL(`Manual transfer needed: ${info.manual} cols → see Section 3 of Summary tab`,'var(--gray)');
    addL('Done! 🎉','var(--success)');setP(100,'Done!');
    document.getElementById('s1').textContent=info.rows||0;
    document.getElementById('s2').textContent=info.mapped||0;
    document.getElementById('s3').textContent=info.comments||0;
    document.getElementById('s4').textContent=(info.not_found||0)+'+'+(info.manual||0);
    document.getElementById('stats').style.display='grid';
    const blob=await res.blob(),url=URL.createObjectURL(blob);
    const dlBtn=document.getElementById('dlBtn');
    dlBtn.style.display='block';
    dlBtn.onclick=()=>{const a=document.createElement('a');a.href=url;a.download=info.filename||'Beta_Converted.xlsm';a.click();};
  }catch(err){addL('ERROR: '+err.message,'var(--gold)');setP(0,'Error!');}
  btn.disabled=false;
};
</script>
</body>
</html>'''

@app.route('/')
def index(): return HTML

@app.route('/convert', methods=['POST'])
def do_convert():
    try:
        std_f  = request.files.get('standard')
        beta_f = request.files.get('beta')
        mp     = request.form.get('marketplace','US')
        std_hdr  = int(request.form.get('std_hdr',3))
        std_data = int(request.form.get('std_data',4))
        beta_hdr = int(request.form.get('beta_hdr',5))
        beta_data= int(request.form.get('beta_data',7))
        if not std_f or not beta_f: return jsonify({'error':'Both files required'}),400
        
        out_buf, rows, mapped, comments, colors, mapped_count, not_found, manual_count = convert(
            std_f.read(), beta_f.read(), mp, std_hdr, std_data, beta_hdr, beta_data)
        
        today = datetime.now().strftime('%d-%m-%Y')
        filename = f'Beta_Converted_{mp}_{today}.xlsm'
        stats = json.dumps({'rows':rows,'mapped':mapped,'comments':comments,'colors':colors,'not_found':not_found,'manual':manual_count,'filename':filename})
        response = send_file(out_buf, mimetype='application/vnd.ms-excel.sheet.macroEnabled.12', as_attachment=True, download_name=filename)
        response.headers['X-Stats'] = stats
        return response
    except Exception as e:
        import traceback
        return jsonify({'error':str(e),'trace':traceback.format_exc()}),500

if __name__=='__main__':
    print("\n"+"="*52)
    print("  ⚡ NEXUS Standard → Beta Converter  v2.0")
    print("  Data · Comments · Colors · Summary Tab")
    print("="*52)
    print("\n  Open: http://localhost:5000")
    print("  Stop: Ctrl+C\n"+"="*52+"\n")
    app.run(debug=False, port=5000)
